#!/usr/bin/env python3
"""
학사 시간표 엑셀 -> src/lib/timetableData.ts 생성기.

사용법:
    python scripts/import_timetable.py "<시간표.xlsx>" [출력경로]

시간표는 개정판이 계속 나오므로(파일명의 날짜 접미사) 재실행 가능하게 유지한다.
알 수 없는 과목이 하나라도 나오면 에러로 중단한다 — 미분류 항목이 조용히
섞여 들어가면 점수 계산이 틀어지기 때문이다.
"""
from __future__ import annotations

import datetime
import io
import json
import re
import sys
from pathlib import Path

import openpyxl

HOLIDAYS = {"추석", "한글날", "대체휴일", "성탄절", "신정", "개천절", "현충일"}

# 의료정보학은 다른 과목 블록과 나란히 진행되어 주(週) 라벨로는 구분되지 않는다.
MEDICAL_INFORMATICS_KEYWORDS = [
    "의료정보",
    "병원정보시스템",
    "진료의사결정시스템",
    "컴퓨터 기반 의학교육",
    "의료데이터베이스",
    "전자의무기록",
]

COURSE_ALIASES = {
    "알레르기-류마티스": "알레르기-류마티스학",
    "임상표현2(기침, 저혈압)": "임상표현2",
    "근골격계학": "근골격학",
}

# src/lib/courses.ts 와 반드시 일치해야 한다.
KNOWN_COURSES = {
    "혈액종양학": "major",
    "순환기학": "major",
    "호흡기학": "major",
    "생식의학": "major",
    "알레르기-류마티스학": "major",
    "내분비학": "major",
    "신경계학": "major",
    "감각계학": "major",
    "근골격학": "major",
    "PBL3": "minor",
    "법의학": "minor",
    "발열": "minor",
    "의료정보학": "minor",
    "임상표현2": "minor",
    "공휴일": "minor",
}


def norm(value) -> str | None:
    if value is None:
        return None
    text = re.sub(r"\s+", " ", str(value).replace("\n", " ")).strip()
    return text or None


def normalize_course(name: str) -> str:
    name = name.strip()
    name = COURSE_ALIASES.get(name, name)
    # "임상표현2(기침, 저혈압)" 처럼 괄호 설명이 붙은 블록명 정리
    stripped = re.sub(r"\(.*\)", "", name).strip()
    return COURSE_ALIASES.get(stripped, stripped or name)


def classify_entry(raw: str) -> str:
    if raw in HOLIDAYS:
        return "holiday"
    if "KAMC" in raw:
        return "exam"
    if re.search(r"(총괄평가|피드백|과정형성평가)", raw):
        return "exam"
    return "lecture"


def resolve_course(raw: str, block: str | None) -> str | None:
    """엑셀 셀 내용과 그 주의 과목 블록으로부터 과목명을 결정한다."""
    if raw in HOLIDAYS:
        return "공휴일"
    if raw.startswith("PBL"):
        return "PBL3"
    if raw == "법의학":
        return "법의학"
    if raw.startswith("발열"):
        return "발열"
    if any(keyword in raw for keyword in MEDICAL_INFORMATICS_KEYWORDS):
        return "의료정보학"
    if raw.startswith("과정형성평가"):
        return normalize_course(block) if block else None
    matched = re.match(r"^(.+?)\s*(총괄평가|피드백)", raw)
    if matched:
        return normalize_course(matched.group(1))
    if block:
        return normalize_course(block)
    return None


def split_topic_professor(raw: str) -> tuple[str, str]:
    """'종양의 이해(한미아)' -> ('종양의 이해', '한미아')"""
    matched = re.match(r"^(.*?)\s*\(([^()]*)\)\s*$", raw)
    if matched and matched.group(2).strip():
        return matched.group(1).strip(), matched.group(2).strip()
    return raw.strip(), ""


def build_merge_maps(sheet):
    anchor_span: dict[tuple[int, int], tuple[int, int]] = {}
    member_anchor: dict[tuple[int, int], tuple[int, int]] = {}
    for rng in sheet.merged_cells.ranges:
        anchor = (rng.min_row, rng.min_col)
        anchor_span[anchor] = (
            rng.max_row - rng.min_row + 1,
            rng.max_col - rng.min_col + 1,
        )
        for row in range(rng.min_row, rng.max_row + 1):
            for col in range(rng.min_col, rng.max_col + 1):
                member_anchor[(row, col)] = anchor
    return anchor_span, member_anchor


def load_split_titles(wb) -> set[str]:
    """<설정> 시트 T열(학습부 분할 강의명) — 이 강의는 학습부 2팀(4명)으로 배정된다."""
    if "설정" not in wb.sheetnames:
        return set()
    settings = wb["설정"]
    titles: set[str] = set()
    for row in range(3, settings.max_row + 1):
        value = norm(settings.cell(row, 20).value)  # T열
        if value:
            titles.add(value)
    return titles


def parse(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = wb["시간표"] if "시간표" in wb.sheetnames else wb.active
    split_titles = load_split_titles(wb)
    anchor_span, member_anchor = build_merge_maps(sheet)

    week_header_rows = [
        row
        for row in range(1, sheet.max_row + 1)
        if isinstance(sheet.cell(row, 3).value, datetime.datetime)
    ]
    week_header_rows.append(sheet.max_row + 1)

    records: list[dict] = []
    unresolved: list[str] = []
    previous_block: str | None = None

    for index in range(len(week_header_rows) - 1):
        header_row = week_header_rows[index]
        end_row = week_header_rows[index + 1]
        week_label = norm(sheet.cell(header_row, 1).value)

        raw_dates: dict[int, datetime.date] = {
            col: value.value.date()
            for col in range(3, 9)
            if isinstance((value := sheet.cell(header_row, col)).value, datetime.datetime)
        }
        if not raw_dates:
            continue

        # 원본의 연도가 1년 밀려 있다(2025-08-31은 일요일이고 2026-08-31이 월요일).
        # 한 주가 연말을 넘기면(12/28~1/2) 같은 주 안에 두 연도가 섞이므로,
        # 그 주의 월요일 기준으로 보정 폭을 정해 주 전체에 똑같이 적용한다.
        monday = raw_dates[min(raw_dates)]
        year_shift = 1 if monday.year <= 2025 else 0
        dates = {
            col: date.replace(year=date.year + year_shift) for col, date in raw_dates.items()
        }

        corrected_monday = dates[min(dates)]
        if corrected_monday.weekday() != 0:
            raise SystemExit(
                f"{week_label}: 보정 후 첫 날짜 {corrected_monday}가 월요일이 아닙니다."
            )

        # 점심행(13:00-14:00)의 첫 비어있지 않은 셀 = 그 주의 과목 블록.
        block: str | None = None
        for row in range(header_row + 1, end_row):
            label = norm(sheet.cell(row, 2).value)
            if label and "13:00-14:00" in label:
                for col in range(3, 9):
                    value = norm(sheet.cell(row, col).value)
                    if value:
                        block = value
                        break
                break
        if block is None:
            block = previous_block  # 17주처럼 라벨이 빠진 주는 직전 주를 승계
        else:
            previous_block = block

        for row in range(header_row + 1, end_row):
            period_no = sheet.cell(row, 1).value
            time_label = norm(sheet.cell(row, 2).value)
            if not isinstance(period_no, (int, float)):
                continue
            matched = re.match(r"\(?(\d{2}):(\d{2})-", time_label or "")
            if not matched:
                continue
            start_hour, start_min = int(matched.group(1)), int(matched.group(2))

            for col in range(3, 9):
                if member_anchor.get((row, col), (row, col)) != (row, col):
                    continue  # 병합 영역의 앵커가 아닌 칸
                raw = norm(sheet.cell(row, col).value)
                if not raw or col not in dates:
                    continue
                row_span, col_span = anchor_span.get((row, col), (1, 1))
                if col_span > 1:
                    continue  # 여러 요일에 걸친 배너(주 과목명 등)

                course = resolve_course(raw, block)
                if course is None or course not in KNOWN_COURSES:
                    unresolved.append(f"{dates[col]} {raw!r} (block={block!r})")
                    continue

                entry_type = classify_entry(raw)
                topic, professor = split_topic_professor(raw)
                end_hour = start_hour + row_span

                records.append(
                    {
                        "week": week_label,
                        "date": dates[col].isoformat(),
                        "order": int(period_no),
                        "period": f"{int(period_no)}교시"
                        if row_span == 1
                        else f"{int(period_no)}~{int(period_no) + row_span - 1}교시",
                        "subject": course,
                        "topic": topic,
                        "professor": professor,
                        "subjectType": KNOWN_COURSES[course],
                        "durationHours": row_span,
                        "startTime": f"{start_hour:02d}:{start_min:02d}",
                        "endTime": f"{end_hour:02d}:{start_min:02d}",
                        "entryType": entry_type,
                        "assignable": entry_type == "lecture",
                        "split": raw in split_titles,
                    }
                )

    if unresolved:
        raise SystemExit(
            "과목을 결정하지 못한 항목이 있습니다:\n  " + "\n  ".join(unresolved)
        )

    records.sort(key=lambda r: (r["date"], r["order"]))

    # <설정> 시트에 "학습부 분할"로 지정된 강의는 학습부 2팀(4명)이 배정되므로
    # 강의 항목 자체를 2개로 복제한다 (팀마다 별도 초안/검안 배정이 붙는다).
    expanded: list[dict] = []
    for record in records:
        if record.pop("split") and record["assignable"]:
            for team in (1, 2):
                clone = dict(record)
                clone["topic"] = f"{record['topic']} ({team}팀 배정)"
                expanded.append(clone)
        else:
            expanded.append(record)
    records = expanded

    # 과목별 학습부 회차 번호 (배정 대상만 센다)
    counters: dict[str, int] = {}
    for record in records:
        if record["assignable"]:
            counters[record["subject"]] = counters.get(record["subject"], 0) + 1
            record["sessionNumber"] = str(counters[record["subject"]])

    return records


def to_typescript(records: list[dict]) -> str:
    lines = [
        "// 자동 생성 파일 — 직접 수정하지 마세요.",
        "// 생성: python scripts/import_timetable.py <시간표.xlsx>",
        'import { Lecture } from "./types";',
        "",
        "type SeedLecture = Omit<Lecture, \"id\" | \"status\">;",
        "",
        "export const TIMETABLE_LECTURES: SeedLecture[] = [",
    ]
    for record in records:
        entry = {
            "date": record["date"],
            "period": record["period"],
            "order": record["order"],
            "subject": record["subject"],
            "topic": record["topic"],
            "professor": record["professor"],
            "subjectType": record["subjectType"],
            "durationHours": record["durationHours"],
            "startTime": record["startTime"],
            "endTime": record["endTime"],
            "entryType": record["entryType"],
            "assignable": record["assignable"],
        }
        if "sessionNumber" in record:
            entry["sessionNumber"] = record["sessionNumber"]
        lines.append("  " + json.dumps(entry, ensure_ascii=False) + ",")
    lines.append("];")
    lines.append("")
    return "\n".join(lines)


def main() -> None:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    if len(sys.argv) < 2:
        raise SystemExit(__doc__)
    source = Path(sys.argv[1])
    target = Path(sys.argv[2]) if len(sys.argv) > 2 else Path("src/lib/timetableData.ts")

    records = parse(source)

    with io.open(target, "w", encoding="utf-8", newline="\n") as handle:
        handle.write(to_typescript(records))

    counts: dict[str, int] = {}
    for record in records:
        counts[record["subject"]] = counts.get(record["subject"], 0) + 1
    assignable = sum(1 for r in records if r["assignable"])

    print(f"{len(records)}건 파싱 -> {target}")
    print(f"  배정 대상(학습부): {assignable}건 / 제외(평가·공휴일): {len(records) - assignable}건")
    print(f"  기간: {records[0]['date']} ~ {records[-1]['date']}")
    for name, count in sorted(counts.items(), key=lambda kv: -kv[1]):
        print(f"  {count:4d}  {name}")


if __name__ == "__main__":
    main()
